#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@File    :   excel_fillin.py
@Time    :   2020/04/25 10:32:48
@Author  :   Jeffrey Wang
@Version :   1.0
@Contact :   shwangjj@163.com
@Desc    :   excel 内容填充示例

openpyxl库可以读写xlsx格式的文件，对于xls旧格式的文件只能用xlrd读，xlwt写来完成了

'''
import xlrd
from xlutils.copy import copy
from openpyxl import load_workbook
# from openpyxl import Workbook
# from openpyxl.chart import BarChart, Series, Reference, BarChart3D
# from openpyxl.styles import Color, Font, Alignment
# from openpyxl.styles.colors import BLUE, RED, GREEN, YELLOW


def change_by_xlrd():
    excel_file = r"src/wdbd/codepool/excel/已有内容.xls"
    output_file = r"src/wdbd/codepool/excel/改写后内容.xls"

    rb = xlrd.open_workbook(excel_file, formatting_info=True)
    wb = copy(rb)       # 利用xlutils.copy下的copy函数复制
    ws = wb.get_sheet(0)
    ws.write(4 - 1, 3 - 1, '添加内容 4,3 ')  # 第4行，第3列
    wb.save(output_file)  # 保存文件，由于save只能保存为xls格式文件
    print('更新后文件是：' + output_file)


def change_by_openpyxl():
    excel_file = r"src/wdbd/codepool/excel/已有内容openpyxl.xlsx"
    output_file = r"src/wdbd/codepool/excel/改写后内容openpyxl.xlsx"
    wb = load_workbook(excel_file)
    ws = wb["Sheet1"]
    print(ws)
    ws["C4"] = "ABC"

    wb.save(output_file)
    print('更新后文件是：' + output_file)


if __name__ == "__main__":
    change_by_openpyxl()
