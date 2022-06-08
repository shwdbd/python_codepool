#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@File    :   jinja_excel.py
@Time    :   2022/01/28 16:01:42
@Author  :   Jeffrey Wang
@Version :   1.0
@Contact :   shwangjj@163.com
@Desc    :   None

template = jxls.Template('xxx.xlsx', 'sheet_name')
template.render(data={}, cell='A3')
template.render(data=df, table='A3')

data = {
    '余额': 12345,
    '明细表': df
}
data_map = {
    'Sheet1': [
        '余额': {cell='A3'},
        '明细表': {table='A3'}
    ]
}
template = jxls.Template('xxx.xlsx', 'sheet_name')
template.render(data, map)

'''
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
# FIXME 单行表，写入
# FIXME xy的写入


class Template:
    """Excel模板渲染器"""

    def __init__(self, excel_path: str):
        try:
            self.file_path = excel_path
            self.wb = load_workbook(filename=excel_path)
        except Exception as err:
            raise IOError(str(err))

    def _set_table(self, sheet, df, cell_id):
        """ 填充表格 """
        row0 = cell_id[0]
        col0 = cell_id[1]
        drow = dcol = 0
        for row in dataframe_to_rows(df.T, index=False, header=False):
            drow = 0
            # print(row)
            for c in row:
                # print('v={v}   x={x},y={y}'.format(v=c, x=x0+dx, y=y0+dy))
                sheet.cell((row0 + drow), (col0 + dcol)).value = c
                drow += 1
            dcol += 1

    def render(self, data: dict, cells: dict, sheet_name: str = None):
        """
        往单一数据单元格 渲染数据

        Example:
        data = {"Name": "Mike", "Age": 12}
        cells = {"Name": "B2", "Age": "C2"}
        Template.render(data, cells)

        """
        if not sheet_name:
            sheet = self.wb.active
        else:
            sheet = self.wb[sheet_name]

        for key in cells.keys():
            cell_id = cells.get(key)
            value = data.get(key)

            if isinstance(value, pd.DataFrame):
                # 表格渲染，cell_id是表头位置
                self._set_table(sheet, value, cell_id)
                # TODO 处理B3这种情况
            else:
                # 单一单元格渲染
                if isinstance(cell_id, str):    # 标签定位
                    sheet[cell_id].value = value
                elif isinstance(cell_id, tuple):    # xy定位
                    sheet.cell(cell_id[0], cell_id[1]).value = value

    # def render_table(self, data: pd.DataFrame, table_header, sheet_name: str)

    def write_excel(self, save_as_path: str = None):
        """将结果写入Excel文件

        Args:
            save_as_path (str, optional): 另存为文件. Defaults to None.
        """
        if not save_as_path:
            self.wb.save(filename=self.file_path)
        else:
            self.wb.save(filename=save_as_path)


if __name__ == "__main__":
    # file_path = 'C:/00 temp/一个单元格模板.xlsx'
    # t = Template(file_path)
    # data = {"Name": "Mike", "Age": 12}
    # cells = {"Name": "B2", "Age": (2, 5)}
    # t.render(data, cells)
    # t.write_excel()

    # 表渲染
    # data = {
    #     'id': range(1, 10),
    #     'value': range(1, 10)
    # }
    # df = pd.DataFrame(data)
    df = pd.DataFrame({"AAA": ["A", "B", "C"], "BBB": [1, 2, 3]})
    print(df)
    file_path = 'C:/00 temp/一个单元格模板.xlsx'
    t = Template(file_path)
    # data = {"表1": df[df["AAA"] == 'A']}
    data = {"表1": df}
    # cells = {"表1": "C3"}
    cells = {"表1": (2, 4)}
    t.render(data, cells)
    t.write_excel()
